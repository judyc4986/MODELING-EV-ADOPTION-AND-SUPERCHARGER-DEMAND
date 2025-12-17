import os
import math
from flask import Flask, render_template, request, send_from_directory, flash, redirect, url_for
from openpyxl import load_workbook

# =========================================================
# APP SETUP
# =========================================================
app = Flask(__name__)
app.secret_key = "random_secret_string"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FORMULA_PATH = os.path.join(BASE_DIR, "static", "data", "equation.xlsx")
SC_SUMMARY_PATH = os.path.join(BASE_DIR, "static", "data", "supercharger_by_county_summary.xlsx")

CHARTS_DIR = os.path.join(BASE_DIR, "static", "charts")
MAPS_DIR = os.path.join(BASE_DIR, "static", "maps")

# =========================================================
# COUNTY NORMALIZER (CRITICAL)
# =========================================================
def normalize_county(name):
    if not name:
        return ""
    return (
        str(name)
        .lower()
        .replace(" county", "")
        .strip()
    )

# =========================================================
# LOAD EQUATION FILE
# =========================================================
if not os.path.exists(FORMULA_PATH):
    raise FileNotFoundError(f"Missing: {FORMULA_PATH}")

wb = load_workbook(FORMULA_PATH)
ws = wb.active

columns = [c.value for c in ws[1]]

formula_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    r = dict(zip(columns, row))
    r["County_clean"] = normalize_county(r["County"])
    formula_rows.append(r)

# =========================================================
# LOAD EXISTING SUPERCARGERS
# =========================================================
if not os.path.exists(SC_SUMMARY_PATH):
    raise FileNotFoundError(f"Missing: {SC_SUMMARY_PATH}")

wb_sc = load_workbook(SC_SUMMARY_PATH)
ws_sc = wb_sc.active

sc_columns = [c.value for c in ws_sc[1]]

supercharger_summary = []
for row in ws_sc.iter_rows(min_row=2, values_only=True):
    r = dict(zip(sc_columns, row))
    r["County_clean"] = normalize_county(r["County"])
    supercharger_summary.append(r)

# =========================================================
# SAFE FORMULA EVALUATOR (SUPPORTS LOGISTIC)
# =========================================================
def evaluate_formula(formula_str, x):
    if not isinstance(formula_str, str):
        return None

    parts = formula_str.split("=")
    if len(parts) < 2:
        return None

    expr = parts[1].strip().replace("^", "**")

    def sigmoid(z):
        try:
            return 1 / (1 + math.exp(-z))
        except OverflowError:
            return 0.0 if z < 0 else 1.0

    allowed = {
        "x": x,
        "L": 1.0,
        "sigmoid": sigmoid,
        "exp": math.exp,
        "log": math.log,
        "sqrt": math.sqrt
    }

    try:
        return float(eval(expr, {"__builtins__": {}}, allowed))
    except Exception as e:
        print("Eval error:", e)
        print("Formula:", formula_str)
        return None

# =========================================================
# IMAGE MATCHER
# =========================================================
def find_image_for_county(directory, county):
    if not os.path.isdir(directory):
        return None

    target = normalize_county(county).replace(" ", "_")

    for f in os.listdir(directory):
        if f.lower().endswith(".png") and target in f.lower():
            return f
    return None

# =========================================================
# STATIC ROUTES
# =========================================================
@app.route("/chart/<path:filename>")
def serve_chart(filename):
    return send_from_directory(CHARTS_DIR, filename)

@app.route("/map/<path:filename>")
def serve_map(filename):
    return send_from_directory(MAPS_DIR, filename)

# =========================================================
# MAIN ROUTE
# =========================================================
@app.route("/", methods=["GET", "POST"])
def index():

    result = None
    county_display = None
    chart_filename = None
    map_filename = None
    existing_sc = None

    if request.method == "POST":
        county_input = request.form.get("county", "")
        year_input = request.form.get("year", "")

        county_clean = normalize_county(county_input)

        if not county_clean:
            flash("Please enter a county name.")
            return redirect(url_for("index"))

        try:
            year = int(year_input)
            if year < 2024 or year > 2050:
                raise ValueError
        except Exception:
            flash("Year must be between 2024 and 2050.")
            return redirect(url_for("index"))

        matches = [r for r in formula_rows if r["County_clean"] == county_clean]
        if not matches:
            flash(f"County '{county_input}' not found.")
            return redirect(url_for("index"))

        row = matches[0]
        county_display = row["County"]

        # Existing superchargers
        sc_match = [r for r in supercharger_summary if r["County_clean"] == county_clean]
        if sc_match:
            existing_sc = sc_match[0].get("Supercharger_Count")

        # Forecast calculations
        sc_formula = row.get("Supercharger_Equation")
        adopt_formula = row.get("Adoption_Equation")
        population = row.get("Population")

        sc_forecast = evaluate_formula(sc_formula, year)
        adopt_rate = evaluate_formula(adopt_formula, year)

        if sc_forecast is None or adopt_rate is None or population is None:
            flash("Could not evaluate equations for this county.")
            return redirect(url_for("index"))

        adopt_rate = max(0.0, min(adopt_rate, 1.0))
        evs = adopt_rate * population

        result = {
            "year": year,
            "sc_forecast": sc_forecast,
            "adopt_rate": adopt_rate,
            "evs": evs
        }

        chart_filename = find_image_for_county(CHARTS_DIR, county_display)
        map_filename = find_image_for_county(MAPS_DIR, county_display)

    return render_template(
        "index.html",
        result=result,
        county=county_display,
        chart_filename=chart_filename,
        map_filename=map_filename,
        existing_sc=existing_sc
    )

# =========================================================
# RUN LOCALLY
# =========================================================
if __name__ == "__main__":
    app.run(debug=True)
