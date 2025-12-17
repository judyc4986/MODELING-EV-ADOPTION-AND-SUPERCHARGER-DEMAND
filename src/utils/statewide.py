from flask import Flask, render_template, request
import openpyxl
import os

app = Flask(__name__)

# =========================================================
# Load statewide supercharger count WITHOUT pandas
# =========================================================
def load_statewide_supercharger_count():
    excel_path = "supercharger_by_county_summary.xlsx"  # must be in project root

    if not os.path.exists(excel_path):
        print("⚠ Missing supercharger_by_county_summary.xlsx")
        return None

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        county_col = None
        sc_col = None

        # detect headers
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col).value).strip().lower()
            if header == "county":
                county_col = col
            if "super" in header or "charger" in header:
                sc_col = col

        if not county_col or not sc_col:
            print("⚠ Required columns not found")
            return None

        # find TOTAL row
        for row in range(2, ws.max_row + 1):
            county_val = str(ws.cell(row=row, column=county_col).value).strip().lower()
            if county_val == "total":
                return int(ws.cell(row=row, column=sc_col).value)

        print("⚠ Row 'Total' not found")
        return None

    except Exception as e:
        print("⚠ Error loading statewide count:", e)
        return None


STATEWIDE_SC = load_statewide_supercharger_count()

# =========================================================
# Constants
# =========================================================
TOTAL_EV_BASE = 3_777_493

# =========================================================
# Statewide forecast by YEAR
# =========================================================
def statewide_forecast_by_year(year):
    t = year - 2024

    # Supercharger points
    superchargers = (
        1.428264e-01 * (t ** 3)
        - 6.879041e+00 * (t ** 2)
        + 1.325427e+02 * t
        + 5.000000e+01
    )

    # Adoption rate
    adoption = (
        -5.999758e-04 * (t ** 2)
        + 4.740971e-02 * t
        + 5.271194e-02
    )

    # guardrails
    adoption = max(0, min(adoption, 1))

    ev_reg = adoption * TOTAL_EV_BASE

    return superchargers, adoption, ev_reg


# =========================================================
# Routes
# =========================================================
@app.route("/", methods=["GET", "POST"])
def index():
    result = None
    error = None

    if request.method == "POST":
        try:
            year = int(request.form.get("year"))

            if year < 2024 or year > 2050:
                raise ValueError

            sc, adopt, ev = statewide_forecast_by_year(year)

            result = {
                "year": year,
                "sc_str": f"{sc:,.0f}",
                "adopt_pct_str": f"{adopt * 100:.2f}",
                "ev_reg_str": f"{ev:,.0f}",
            }

        except Exception:
            error = "Please enter a valid year between 2024 and 2050."

    return render_template(
        "index.html",
        result=result,
        error=error,
        statewide_sc=STATEWIDE_SC,
    )


if __name__ == "__main__":
    app.run(debug=True)
